import time

import openpyxl
import datetime
import re

from datetime import datetime, timedelta

number_row_1 = 0
date_old = datetime
date_old_min = datetime
name_sheet = ''
ws = ''
wb = ''
ws_1 = ''
path = None
path_file = None
m = 0
temp = ''
date_check = ''
date_start_input = ''
date_end_input = ''
min_row = 0
max_row = 0
flag_finish = 0
d1 = None
d2 = None
d3 = None


def today_sheet():
    global ws_1, wb
    today = str(datetime.today().strftime("Данные от %d.%m.%Y | %H-%M-%S"))
    wb.create_sheet(index=1, title=today)
    ws_1 = wb[today]

    for row in ws['A1':'BK1']:
        for cell in row:
            header_number = cell.column
            header_info = cell.value
            c1 = ws_1.cell(row=1, column=header_number)
            c1.value = header_info


def excel_create(path_file):
    global wb, ws, path, name_sheet
    path = fr'{path_file}'
    wb = openpyxl.load_workbook(filename=path, read_only=False)
    ws = wb['Page 1']

    today_sheet()

    name_sheet = str(ws_1).replace('<', '').replace('>', '').replace('Worksheet', '')


def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days)):
        yield start_date + timedelta(n)


def remove(ws_1):
    global k, m, max_row

    # items = list(range(1, max_row))

    while m != max_row:
        m = m + 1
        for row in ws_1.iter_rows(min_row=1, max_row=max_row):
            if not any(cell.value for cell in row):
                ws_1.delete_rows(row[0].row, 1)
                print(f'Строка №{m} удалена... Осталось: {max_row - m}')

                remove(ws_1)

                return
    m = 0
    pass


def znach(date_check):
    date_temple = datetime.strptime(date_check, '%d.%m.%Y')

    date_min = date_temple.strftime("%d.%m.%y")
    date_max = date_temple.strftime("%d.%m.%Y")

    number_row_1 = 0

    for row in ws.rows:
        number_row_1 += 1
        for cell in row:
            if re.match(fr'{date_max}', str(cell.value)):
                for cell in row:
                    col_number = cell.column
                    cell_info = cell.value

                    c1 = ws_1.cell(row=number_row_1, column=col_number)
                    c1.value = cell_info
                print(f'Строка №{number_row_1} успешно записана!')

            if re.match(fr'{date_min}', str(cell.value)):
                for cell in row:
                    col_number = cell.column
                    cell_info = cell.value

                    c1 = ws_1.cell(row=number_row_1, column=col_number)
                    c1.value = cell_info
                print(f'Строка №{number_row_1} успешно записана!')
    finish()


def rage(date_start_input, date_end_input):
    global date_old
    start_date = datetime.strptime(date_start_input, '%d.%m.%Y')
    end_date_temp = datetime.strptime(date_end_input, '%d.%m.%Y')
    end_date = end_date_temp + timedelta(days=1)

    for single_date in daterange(start_date, end_date):
        date_old = single_date.strftime("%d.%m.%Y")
        date_old_min = single_date.strftime("%d.%m.%y")
        diapazon_date(date_old_min, date_old)
    finish()


def diapazon_date(date_temp_min, date_temp):
    number_row_diap = 0
    for row in ws.rows:
        number_row_diap += 1
        for cell in row:
            if re.match(fr'{date_temp}', str(cell.value)):
                for cell in row:
                    col_number = cell.column
                    cell_info = cell.value

                    c1 = ws_1.cell(row=number_row_diap, column=col_number)
                    c1.value = cell_info
                print(f'Строка №{number_row_diap} успешно записана!')

            if re.match(fr'{date_temp_min}', str(cell.value)):
                for cell in row:
                    col_number = cell.column
                    cell_info = cell.value

                    c1 = ws_1.cell(row=number_row_diap, column=col_number)
                    c1.value = cell_info
                print(f'Строка №{number_row_diap} успешно записана!')


def check(date_check):
    global flag_finish, d3
    try:
        d3 = datetime.strptime(date_check, '%d.%m.%Y')
        if d3 is not None:
            znach(date_check)
    except:
        if flag_finish != 1:
            print('Вы ввели не корректную дату!')
            start()


def check_diap(date_start_input, date_end_input):
    global flag_finish, d1, d2
    try:
        d1 =  datetime.strptime(date_start_input, '%d.%m.%Y')
        d2 = datetime.strptime(date_end_input, '%d.%m.%Y')
        if d1 is not None and d2 is not None:
            rage(date_start_input, date_end_input)
    except:
        if flag_finish != 1:
            print('Вы ввели не корректную дату!')
            start()


def start():
    global temp, date_check, date_start_input, date_end_input

    temp = input(str('\nВы хотите выполнить поиск по значению (з) или диапазону (д)? \nВведите з/д: '))
    if temp == 'з' or temp == 'З' or temp == 'д' or temp == 'Д':
        if temp == 'з' or temp == 'З':
            date_check = input('Введите дату в формате "01.01.2001": ')
            check(date_check)
        elif temp == 'д' or temp == 'Д':
            date_start_input = input('Введите начальную дату в формате "01.01.2001": ')
            date_end_input = input('Введите конечную дату в формате "01.01.2001": ')
            check_diap(date_start_input, date_end_input)
    else:
        print('Вы ввели не корректный ответ!')
        start()


def create_path():
    global path_file, flag_finish
    if flag_finish != 1:
        if path_file is None:
            path_file = input(str('Введите путь до файла Excel: '))
            try:
                excel_create(path_file)
                start()
            except:
                if flag_finish != 1:
                    print('Не удается найти файл Excel! Проверьте путь или уберите пробел в конце пути')
                    path_file = None
                    create_path()
        else:
            exit(0)


def answer():
    global flag_finish
    flag_finish = 1
    finish_temp = input(str("Программа выполнила работу! \nХотите воспользоваться программой еще раз?: д/н "))
    if finish_temp == 'д' or finish_temp == 'Д' or finish_temp == 'н' or finish_temp == 'Н':
        if finish_temp == 'д' or finish_temp == 'Д':
            today_sheet()
            flag_finish = 0
            start()
        elif finish_temp == 'н' or finish_temp == 'Н':
            print('\n Хорошего дня! :)')
            exit(0)
    else:
        print('Вы ввели не корректный ответ!')
        answer()


def finish():
    global min_row, max_row, ws_1, temp, date_check, date_start_input, date_end_input, name_sheet
    min_row = ws_1.min_row
    max_row = ws_1.max_row
    remove(ws_1)

    wb.save(path)

    if temp == 'з' or temp == 'З':
        print(f"\n\nВаш запрос: данные от {date_check}")
    elif temp == 'д' or temp == 'Д':
        print(f"\n\nВаш запрос: данные с {date_start_input} по {date_end_input}")

    print("\nМаксимум строк до удаления:", max_row)
    print("Максимум строк после удаления:", ws_1.max_row)

    print(f"\nНовый лист с именем {name_sheet} создан! \n")
    answer()


if __name__ == "__main__":
    if path_file is None:
        create_path()
        # path_file = input(str('Введите путь до файла Excel: '))
        # excel_create(path_file)
        # start()

# Проверка на корректную дату + пофиксить краш при неверном пути до файла
