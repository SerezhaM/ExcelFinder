import openpyxl
import datetime
import re

import tkinter as tk
import tkinter.font as tkFont
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askopenfilename

from datetime import datetime, timedelta


number_row_1 = 0
date_old = datetime
date_old_min = datetime
name_sheet = ''
ws = ''
wb = ''
ws_1 = ''
path = None
path_file = None
m = 0
temp = ''
date_save = None
date_start_input = ''
date_end_input = ''
min_row = 0
max_row = 0
flag_finish = 0
d1 = None
d2 = None
d3 = None



label_start = ''
button_file = ''
label_file = ''
label_choose_date = ''
btn_znach = ''
btn_diap = ''
entry_znach = ''
entry_diap_start = ''
entry_diap_end = ''
label_date_znach = ''
btn_1 = ''
label_info = None
label_lucky = ''
label_remove_info = None
label_remove_lucky = ''
label_1 = None
label_2 = ''
label_3 = ''
btn_restart = ''
btn_exit = ''



def today_sheet():
    global ws_1, wb
    today = str(datetime.today().strftime("Данные от %d.%m.%Y | %H-%M-%S"))
    wb.create_sheet(index=1, title=today)
    ws_1 = wb[today]

    for row in ws['A1':'BK1']:
        for cell in row:
            header_number = cell.column
            header_info = cell.value
            c1 = ws_1.cell(row=1, column=header_number)
            c1.value = header_info


def excel_create(path_file):
    global wb, ws, path, name_sheet
    path = fr'{path_file}'
    wb = openpyxl.load_workbook(filename=path, read_only=False)
    ws = wb['Page 1']

    today_sheet()

    name_sheet = str(ws_1).replace('<', '').replace('>', '').replace('Worksheet', '')


def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days)):
        yield start_date + timedelta(n)


def remove(ws_1):
    global k, m, max_row, label_remove_info, label_remove_lucky, message_remove

    # label_remove_info = tk.Label(root, textvariable=message_remove, font=fontExample, bg='white', fg='black')
    label_remove_info = tk.Label(text='', font=fontExample, bg="white", fg='black')
    label_remove_info.place(relx=0.4, rely=0.2, anchor="center")

    while m != max_row:
        m = m + 1
        for row in ws_1.iter_rows(min_row=1, max_row=max_row):
            if not any(cell.value for cell in row):
                ws_1.delete_rows(row[0].row, 1)

                # message_remove.set(f'Строка №{m} удалена... Осталось: {max_row - m}')
                # root.update_idletasks()
                label_remove_info.config(text=f'Строка №{m} удалена... Осталось: {max_row - m}')
                root.update()

                remove(ws_1)

                return
    m = 0
    label_remove_lucky = tk.Label(root, text='Успешно!', font=fontExample, fg='green', bg='white')
    label_remove_lucky.place(relx=0.85, rely=0.2, anchor="center")
    pass


def znach(date_check):
    global label_info, label_lucky

    btn_1.place_forget()
    label_date_znach.place_forget()
    entry_znach.place_forget()

    date_temple = datetime.strptime(date_check, '%d.%m.%Y')

    date_min = date_temple.strftime("%d.%m.%y")
    date_max = date_temple.strftime("%d.%m.%Y")

    number_row_1 = 0

    label_info = tk.Label(text='text_info', font=fontExample, bg="white", fg='black')
    label_info.place(relx=0.43, rely=0.1, anchor="center")

    for row in ws.rows:
        number_row_1 += 1
        for cell in row:
            if re.match(fr'{date_max}', str(cell.value)):
                for cell in row:
                    col_number = cell.column
                    cell_info = cell.value

                    c1 = ws_1.cell(row=number_row_1, column=col_number)
                    c1.value = cell_info
                label_info.config(text=f'Строка №{number_row_1} успешно записана!')
                root.update()
            if re.match(fr'{date_min}', str(cell.value)):
                for cell in row:
                    col_number = cell.column
                    cell_info = cell.value


                    c1 = ws_1.cell(row=number_row_1, column=col_number)
                    c1.value = cell_info
                label_info.config(text=f'Строка №{number_row_1} успешно записана!')
                root.update()
    label_lucky = tk.Label(root, text='Успешно!', font=fontExample, fg='green', bg='white')
    label_lucky.place(relx=0.85, rely=0.1, anchor="center")
    finish()


def rage(date_start_input, date_end_input):
    global date_old, label_lucky
    start_date = datetime.strptime(date_start_input, '%d.%m.%Y')
    end_date_temp = datetime.strptime(date_end_input, '%d.%m.%Y')
    end_date = end_date_temp + timedelta(days=1)

    for single_date in daterange(start_date, end_date):
        date_old = single_date.strftime("%d.%m.%Y")
        date_old_min = single_date.strftime("%d.%m.%y")
        diapazon_date(date_old_min, date_old)

    label_lucky = tk.Label(root, text='Успешно!', font=fontExample, fg='green', bg='white')
    label_lucky.place(relx=0.85, rely=0.1, anchor="center")
    finish()


def diapazon_date(date_temp_min, date_temp):
    global message_text, btn_2, entry_diap_end, entry_diap_start, label_date_diap_end, label_date_diap_start, label_info, label_lucky

    btn_2.place_forget()
    entry_diap_end.place_forget()
    entry_diap_start.place_forget()
    label_date_diap_end.place_forget()
    label_date_diap_start.place_forget()

    number_row_diap = 0

    label_info = tk.Label(root, textvariable=message_text, font=fontExample, bg='white', fg='black')
    label_info.place(relx=0.43, rely=0.1, anchor="center")

    for row in ws.rows:
        number_row_diap += 1
        for cell in row:
            if re.match(fr'{date_temp}', str(cell.value)):
                for cell in row:
                    col_number = cell.column
                    cell_info = cell.value

                    c1 = ws_1.cell(row=number_row_diap, column=col_number)
                    c1.value = cell_info
                message_text.set(f'Строка №{number_row_diap} успешно записана!')
                root.update_idletasks()
            if re.match(fr'{date_temp_min}', str(cell.value)):
                for cell in row:
                    col_number = cell.column
                    cell_info = cell.value

                    c1 = ws_1.cell(row=number_row_diap, column=col_number)
                    c1.value = cell_info
                message_text.set(f'Строка №{number_row_diap} успешно записана!')
                root.update_idletasks()


def check():
    global flag_finish, d3, entry_znach, date_save
    date_save = entry_znach.get()
    try:
        d3 = datetime.strptime(date_save, '%d.%m.%Y')
        if d3 is not None:
            znach(date_save)
    except:
        if flag_finish != 1 and d3 is None:
            messagebox.showinfo(title='Внимание', message=f'Вы ввели: {date_save}! \n\nДата не корректна!')
            label_date_znach.place_forget()
            entry_znach.place_forget()
            btn_1.place_forget()
            date_znach()


def check_diap():
    global flag_finish, d1, d2, entry_diap_start, entry_diap_end

    date_start_input = entry_diap_start.get()
    date_end_input = entry_diap_end.get()

    try:
        d1 =  datetime.strptime(date_start_input, '%d.%m.%Y')
        d2 = datetime.strptime(date_end_input, '%d.%m.%Y')
        if d1 is not None and d2 is not None:
            rage(date_start_input, date_end_input)
    except:
        if flag_finish != 1:
            if d1 is None and d2 is None:
                messagebox.showinfo(title='Внимание', message=f'Вы ввели начальную дату: "{date_start_input}" \nКонечную дату: "{date_end_input}" \n\nДаты не корректны!')
                label_date_diap_start.place_forget()
                label_date_diap_end.place_forget()
                entry_diap_start.place_forget()
                entry_diap_end.place_forget()
                btn_2.place_forget()
                date_diap()
            if d1 is None:
                messagebox.showinfo(title='Внимание', message=f'Вы ввели начальную дату: "{date_end_input}"! \n\nДата не корректна!')
                label_date_diap_start.place_forget()
                label_date_diap_end.place_forget()
                entry_diap_start.place_forget()
                entry_diap_end.place_forget()
                btn_2.place_forget()
                date_diap()
            if d2 is None:
                messagebox.showinfo(title='Внимание', message=f'Вы ввели конечную дату: "{date_end_input}"! \n\nДата не корректна!')
                label_date_diap_start.place_forget()
                label_date_diap_end.place_forget()
                entry_diap_start.place_forget()
                entry_diap_end.place_forget()
                btn_2.place_forget()
                date_diap()


def date_znach():
    global entry_znach, label_file, label_choose_date, btn_znach, btn_diap, temp, label_date_znach, entry_znach, btn_1
    label_choose_date.place_forget()
    btn_znach.place_forget()
    btn_diap.place_forget()
    label_file.place_forget()

    temp = 'з'

    label_date_znach = tk.Label(root, text='Введите дату в формате "01.01.2001"', font=fontExample, bg='white', fg='black')
    label_date_znach.place(relx=0.35, rely=0.4, anchor="center")

    entry_znach = tk.Entry(root, font=fontExample, width=10)
    entry_znach.place(relx=0.8, rely=0.4, anchor="center")

    btn_1 = tk.Button(root, text='Ввести', command=check, font=fontExample, width=30, bd=0, bg='white', fg='black')

    btn_1.place(relx=0.5, rely=0.5, anchor="center")


def date_diap():
    global label_file, entry_diap_start, entry_diap_end, label_choose_date, btn_znach, btn_diap, temp, btn_2, entry_diap_end, entry_diap_start, label_date_diap_end, label_date_diap_start
    label_choose_date.place_forget()
    label_file.place_forget()
    btn_znach.place_forget()
    btn_diap.place_forget()

    temp = 'д'

    label_date_diap_start = tk.Label(root, text='Введите начальную дату \n в формате "01.01.2001"', font=fontExample, width=20, bg='white', fg='black')
    label_date_diap_start.place(relx=0.37, rely=0.3, anchor="center")

    entry_diap_start = tk.Entry(root, font=fontExample, width=10)
    entry_diap_start.place(relx=0.73, rely=0.3, anchor="center")

    label_date_diap_end = tk.Label(root, text='Введите конечную дату \n в формате "01.01.2001"', font=fontExample, width=20, bg='white', fg='black')
    label_date_diap_end.place(relx=0.37, rely=0.4, anchor="center")

    entry_diap_end = tk.Entry(root, font=fontExample, width=10)
    entry_diap_end.place(relx=0.73, rely=0.4, anchor="center")

    btn_2 = tk.Button(root, text='Ввести',
                         command=check_diap, font=fontExample, width=30, bd=0, bg='white', fg='black')

    btn_2.place(relx=0.5, rely=0.5, anchor="center")


def start():
    global label_start, button_file, label_choose_date, btn_znach, btn_diap, path, label_file

    label_start.place(relx=-100, rely=-100, anchor="center")
    button_file.place(relx=-100, rely=-100, anchor="center")

    label_file = tk.Label(root, text=f'Ваш файл:\n{path}', font=fontExample, bg='white', fg='black')
    label_file.place(relx=0.5, rely=0.3, anchor="center")

    label_choose_date = tk.Label(root, text='Вы хотите выполнить поиск по значению или диапазону?', font=fontExample, bg='white', fg='black')
    label_choose_date.place(relx=0.5, rely=0.43, anchor="center")

    btn_znach = tk.Button(root, text='Значение',
                               command=date_znach, font=fontExample, bd=0, bg='white', fg='black')
    btn_diap = tk.Button(root, text='Диапазон',
                              command=date_diap, font=fontExample, bd=0, bg='white', fg='black')

    btn_znach.place(relx=0.37, rely=0.5, anchor="center")
    btn_diap.place(relx=0.63, rely=0.5, anchor="center")


# def restart():
#     global temp, flag_finish, entry_znach, entry_diap_start, entry_diap_end, label_info, label_lucky, \
#         label_1, label_2, label_3, btn_restart, btn_exit, label_remove_lucky, label_remove_info
#
#     if temp == 'з':
#         entry_znach.destroy()
#         label_1.destroy()
#     else:
#         entry_diap_start.destroy()
#         entry_diap_end.destroy()
#         label_2.destroy()
#
#     label_info.destroy()
#     label_remove_info.destroy()
#
#     # label_info.place(relx=0.1, rely=0.1, anchor="center")
#     # label_remove_info.place_forget()
#
#     label_3.destroy()
#     btn_restart.destroy()
#     btn_exit.destroy()
#     label_lucky.destroy()
#     label_remove_lucky.destroy()
#
#     today_sheet()
#     flag_finish = 0
#     start()


def finish():
    global m, btn_exit, min_row, max_row, ws_1, temp, flag_finish, date_save, date_start_input, date_end_input, name_sheet, label_info, label_lucky, label_remove_info, label_remove_lucky, label_1, label_2, label_3, btn_restart, btn_exit

    min_row = ws_1.min_row
    max_row = ws_1.max_row

    remove(ws_1)
    wb.save(path)

    if temp == 'з':
        label_1 = tk.Label(root, text=f"Ваш запрос: данные от {date_save}", font=fontExample, bg='white', fg='black')
        label_1.place(relx=0.5, rely=0.3, anchor="center")
    elif temp == 'д':
        label_2 = tk.Label(root, text=f"Ваш запрос: \nданные с {date_start_input} по {date_end_input}", font=fontExample, bg='white', fg='black')
        label_2.place(relx=0.5, rely=0.3, anchor="center")

    flag_finish = 1

    label_3 = tk.Label(root, text=f"Максимум строк до удаления: {max_row}\nМаксимум строк после удаления: {ws_1.max_row}\n\nНовый лист с именем: \n{name_sheet} создан! \n\nПрограмма успешно выполнила работу!", font=fontExample, bg='white', fg='black')
    label_3.place(relx=0.5, rely=0.4, anchor="center")

    # btn_restart = tk.Button(root, text='Да', command=restart, font=fontExample)
    btn_exit = tk.Button(root, text='Выход', command=exit, font=fontExample, width=30, bd=0, bg='white', fg='black')

    # btn_restart.place(relx=0.4, rely=0.6, anchor="center")
    btn_exit.place(relx=0.5, rely=0.6, anchor="center")



if __name__ == "__main__":

    root = tk.Tk()

    root.geometry(f"500x500")
    root.title("ExcelFinder")
    root.resizable(False, False)

    root.update_idletasks()
    w, h = root.winfo_width(), root.winfo_height()
    root.geometry(f"+{(root.winfo_screenwidth() - w) // 2}+{(root.winfo_screenheight() - h) // 2}")

    message_text = StringVar()
    message_remove = StringVar()

    fontExample = tkFont.Font(family="Segoe UI", size=16, weight="bold")

    background_image = tk.PhotoImage(file='/Users/sergeymashokha/PycharmProjects/excel_prjct/bg.png')
    background_label = tk.Label(root, image=background_image)
    background_label.place(x=0, y=0, relwidth=1, relheight=1)

    # --------
    label_start = tk.Label(root, text='Выберите файл Excel', font=fontExample, fg='#30C',bg='#6C9')
    label_start.place(relx=0.5, rely=0.43, anchor="center")

    def select_file():
        filetypes = (
            ('Excel files', '*.xlsx'),
            ('Excel files', '*.xlsm'),
            ('Excel files', '*.xltx'),
            ('Excel files', '*.xltm'),
        )

        path_file = askopenfilename(
            title='Выбрать файл',
            initialdir='/',
            filetypes=filetypes)

        if path_file is not None:
            excel_create(path_file)
            start()



    button_file = tk.Button(
        root,
        text='Выбрать файл',
        command=select_file,
        font=fontExample,
        background="#fff",
        foreground="#000",
        bd=0,
    )

    button_file.place(relx=0.5, rely=0.5, anchor="center")

    label_name = tk.Label(root, text='Made by Serezha M', font="Arial 9", fg='#360', bg='#6C9')
    label_name.place(relx=0.9, rely=0.95, anchor="center")

root.mainloop()